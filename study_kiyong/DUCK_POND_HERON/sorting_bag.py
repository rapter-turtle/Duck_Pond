import rospy
import pandas as pd
from std_msgs.msg import Float64MultiArray
from heron_msgs.msg import Drive
from mpc_msgs.msg import MPCTraj, MPC_State, Obs_State
import os
from itertools import chain
from openpyxl import load_workbook

class DataSaver:
    def __init__(self):
        # Initialize ROS node
        rospy.init_node('data_saver', anonymous=True)

        # Subscribers
        self.l1_data_sub = rospy.Subscriber('/L1_data', Float64MultiArray, self.l1_callback)
        self.mpc_traj_sub = rospy.Subscriber('/mpc_traj', Float64MultiArray, self.traj_callback)
        self.mpc_vis_sub = rospy.Subscriber('/mpc_vis', MPCTraj, self.vis_callback)
        self.cmd_drive_sub = rospy.Subscriber('/cmd_drive', Drive, self.cmd_drive_callback)
        self.ekf_state_sub = rospy.Subscriber('/ekf/estimated_state', Float64MultiArray, self.ekf_callback)

        # Initialize variables for the latest data
        self.l1_data = []
        self.mpc_traj = []
        self.mpc_vis = []
        self.cmd_drive_left = None
        self.cmd_drive_right = None
        self.ekf_estimated = []
        
        # Buffer for storing data rows in memory
        self.data_buffer = []
        self.save_interval = 1  # Number of rows to buffer before saving

        # File path and initialization
        self.file_path = 'ros_data.xlsx'
        self.init_excel()
        self.kk = 0

    def init_excel(self):
        """Initialize the Excel file with headers if it doesn't exist."""
        if not os.path.isfile(self.file_path):
            headers = (
                ['Index'] +
                [f'L1_data {i}' for i in range(8)] +
                [f'mpc_traj {i}' for i in range(11)] +
                [f'state x {i}' for i in range(50)] +
                [f'state y {i}' for i in range(50)] +
                [f'ref x {i}' for i in range(50)] +
                [f'ref y {i}' for i in range(50)] +
                ['cmd_drive left', 'cmd_drive right'] +
                [f'ekf_estimated {i}' for i in range(12)]
            )
            pd.DataFrame(columns=headers).to_excel(self.file_path, index=False)

    def l1_callback(self, msg):
        self.l1_data = msg.data

    def traj_callback(self, msg):
        self.mpc_traj = msg.data

    def vis_callback(self, msg):
        # Collect only the required data from `state` and `ref` for up to 50 entries
        state_x_data = [s.x for s in msg.state[:50]]
        state_y_data = [s.y for s in msg.state[:50]]
        ref_x_data = [r.x for r in msg.ref[:50]]
        ref_y_data = [r.y for r in msg.ref[:50]]
        
        # Combine the data into a single list
        self.mpc_vis = state_x_data + state_y_data + ref_x_data + ref_y_data

    def ekf_callback(self, msg):
        self.ekf_estimated = msg.data

    def cmd_drive_callback(self, msg):
        # Update latest /cmd_drive data
        self.cmd_drive_left = msg.left
        self.cmd_drive_right = msg.right

        # Save the collected data to the buffer
        self.save_to_buffer()

    def save_to_buffer(self):
        # Load the existing workbook to determine the next index
        if os.path.exists(self.file_path):
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            next_index = sheet.max_row
        else:
            next_index = 1  # Start with 1 if the file doesn't exist

        # Prepare row data with the index as the first column
        row_data = (
            [next_index] +  # Add the index as the first element
            list(self.l1_data) +
            list(self.mpc_traj) +
            list(self.mpc_vis) +
            [self.cmd_drive_left, self.cmd_drive_right] +
            list(self.ekf_estimated)
        )

        # Append the row data to the buffer
        self.data_buffer.append(row_data)

        # Check if the buffer size has reached the save interval
        if len(self.data_buffer) >= self.save_interval:
            self.save_to_excel()

    def save_to_excel(self):
        # Convert buffer to a DataFrame
        df = pd.DataFrame(self.data_buffer)

        # Load the workbook to append new rows, using openpyxl directly
        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        start_row = sheet.max_row + 1  # Find the next empty row

        # Append each row in the DataFrame to the existing sheet
        for row in df.itertuples(index=False, name=None):
            sheet.append(row)

        # Save the workbook after appending
        workbook.save(self.file_path)

        # Clear the buffer after saving
        self.data_buffer = []
        rospy.loginfo("Buffered data saved to ros_data.xlsx")
        
        self.kk += 1
        print(self.kk) 

if __name__ == '__main__':
    try:
        saver = DataSaver()
        rospy.spin()
    except rospy.ROSInterruptException:
        pass
